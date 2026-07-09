"""
================================================================================
RetailMax Enterprise Data Platform

Module:      report_generator.py
Purpose:     Generates Executive Reports (Markdown CEO Dashboard & OpenPyXL Excel sheet)
Author:      Himanshu Sardana
Copyright:   (c) 2026 RetailMax Corp. All rights reserved.
================================================================================
"""

from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import REPORTS_DIR
from employee import Employee
from exceptions import AnalyticsError
from logging_config import get_logger

# Initialize logger
logger = get_logger("report_generator")


class ExecutiveReportGenerator:
    """Creates formatted reports (Markdown dashboard and styled Excel spreadsheets) for executives."""

    def __init__(self, reports_dir: Path = REPORTS_DIR) -> None:
        self.reports_dir = reports_dir
        self.reports_dir.mkdir(parents=True, exist_ok=True)
        logger.debug(f"ExecutiveReportGenerator initialized writing to: {self.reports_dir}")

    def generate_ceo_markdown_dashboard(
        self, kpis: dict[str, Any], dept_summary: list[dict[str, Any]]
    ) -> Path:
        """Renders an executive-ready Markdown dashboard summary.

        Args:
            kpis: Dictionary of pre-calculated corporate KPIs.
            dept_summary: Department aggregates summary list.

        Returns:
            Path: The generated Markdown report path.
        """
        logger.info("Generating CEO Markdown Dashboard report...")
        report_path = self.reports_dir / "ceo_dashboard.md"

        try:
            content = []
            content.append("# RetailMax Enterprise Data Platform - CEO Executive Dashboard")
            content.append(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

            content.append("## 📈 Global HR Key Performance Indicators")
            content.append("| Metric | Value | Description |")
            content.append("| :--- | :--- | :--- |")
            content.append(
                f"| **Total Headcount** | {kpis['Headcount']} | Active employees across all stores |"
            )
            content.append(
                f"| **Total Monthly Salary Bill** | ₹{kpis['TotalSalaryExpense']:,} | Monthly base payroll budget |"
            )
            content.append(
                f"| **Average Monthly Salary** | ₹{kpis['AverageSalary']:,} | Mean employee monthly payout |"
            )
            content.append(
                f"| **Median Monthly Salary** | ₹{kpis['MedianSalary']:,} | Mid-point salary standard |"
            )
            content.append(
                f"| **Highest Monthly Salary** | ₹{kpis['MaxSalary']:,} | Top earner ceiling |"
            )
            content.append(
                f"| **Lowest Monthly Salary** | ₹{kpis['MinSalary']:,} | Floor level entry salary |"
            )
            content.append(
                f"| **Average Tenure (Years)** | {kpis['AverageTenureYears']} years | Mean employee retention length |"
            )
            content.append(
                f"| **Average Performance Score** | {kpis['AveragePerformanceScore']} / 5.0 | Overall organizational rating |"
            )
            content.append(
                f"| **High Performers Count** | {kpis['HighPerformersCount']} | Rating score of 4 or 5 |"
            )
            content.append(
                f"| **High Performers Ratio** | {kpis['HighPerformersRatio']}% | Percentage of outstanding staff |"
            )
            content.append("\n")

            content.append("## 🏢 Department Breakdown Aggregates")
            content.append(
                "| Department | Headcount | Avg Monthly Salary | Max Salary | Total Monthly Expense |"
            )
            content.append("| :--- | :---: | :---: | :---: | :--- |")
            for dept in dept_summary:
                hc = dept["Headcount"]
                avg_sal = f"₹{dept['AverageSalary']:,}"
                max_sal = f"₹{dept['MaxSalary']:,}"
                tot_sal = f"₹{dept['TotalSalary']:,}"
                content.append(
                    f"| **{dept['Department']}** | {hc} | {avg_sal} | {max_sal} | {tot_sal} |"
                )
            content.append("\n")

            content.append("## 🏆 Top 5 Highest Earning Employees")
            content.append("| ID | Name | Department | Monthly Salary |")
            content.append("| :--- | :--- | :--- | :--- |")
            for emp in kpis["Top5Earners"]:
                sal = f"₹{emp['Salary']:,}"
                content.append(
                    f"| {emp['EmployeeID']} | {emp['Name']} | {emp['Department']} | {sal} |"
                )
            content.append("\n")

            # Link generated charts from charts/ folder for visual embedding in markdown editors
            content.append("## 📊 Visual Analytics & Trends")
            content.append(
                "To view full graphical representations, open the corresponding image files inside the `charts/` folder:\n"
            )
            content.append(
                "1. **Salary Distribution Profile:** ![Salary Distribution](../charts/04_salary_distribution.png)"
            )
            content.append(
                "2. **Headcount by Division:** ![Headcount by Dept](../charts/01_headcount_by_dept.png)"
            )
            content.append(
                "3. **Average Salary by Department:** ![Salary by Dept](../charts/03_average_salary_by_dept.png)"
            )
            content.append(
                "4. **Company Growth Trend:** ![Hiring Trends](../charts/08_hiring_trends_line.png)"
            )
            content.append(
                "5. **Performance vs Salary:** ![Performance vs Salary Scatter](../charts/07_performance_vs_salary_scatter.png)"
            )
            content.append(
                "6. **Statistical Ranges (Boxplot):** ![Salary Ranges Boxplot](../charts/05_salary_by_dept_boxplot.png)"
            )
            content.append(
                "7. **Salary probability distributions:** ![Salary Violin Plot](../charts/15_salary_violin_plot.png)"
            )
            content.append(
                "8. **Core correlation matrices:** ![Correlation Heatmap](../charts/16_correlation_heatmap.png)"
            )

            with open(report_path, "w", encoding="utf-8") as f:
                f.write("\n".join(content))

            logger.info(f"Markdown dashboard generated at: {report_path}")
            return report_path
        except Exception as e:
            raise AnalyticsError(f"Failed to generate Markdown report: {e}") from e

    def generate_excel_workbook(
        self, employees: list[Employee], dept_summary: list[dict[str, Any]]
    ) -> Path:
        """Creates a styled, multi-sheet Excel spreadsheet using OpenPyXL.

        Sheet 1: Employee Directory (Full listing with auto-adjusted widths)
        Sheet 2: Department Summaries

        Args:
            employees: Full list of employee records.
            dept_summary: Department aggregates summary.

        Returns:
            Path: The generated Excel workbook path.
        """
        logger.info("Generating corporate Excel spreadsheet report...")
        excel_path = self.reports_dir / "hr_annual_report.xlsx"

        try:
            wb = openpyxl.Workbook()
            # Remove default sheet
            wb.remove(wb.active)

            # Styling definitions
            font_title = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
            font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            font_data = Font(name="Calibri", size=11)
            font_summary = Font(name="Calibri", size=11, bold=True)

            fill_primary = PatternFill(
                start_color="1F4E79", end_color="1F4E79", fill_type="solid"
            )  # Dark steel blue
            fill_accent = PatternFill(
                start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"
            )  # Ice blue

            thin_border_side = Side(style="thin", color="D3D3D3")
            border_cell = Border(
                left=thin_border_side,
                right=thin_border_side,
                top=thin_border_side,
                bottom=thin_border_side,
            )
            double_bottom_side = Side(style="double", color="000000")
            border_summary = Border(top=thin_border_side, bottom=double_bottom_side)

            # ==============================================================================
            # SHEET 1: EMPLOYEE DIRECTORY
            # ==============================================================================
            ws1 = wb.create_sheet(title="Employee Directory")
            ws1.views.sheetView[0].showGridLines = True

            # Write Title Banner
            ws1.merge_cells("A1:K1")
            title_cell = ws1["A1"]
            title_cell.value = "RetailMax Enterprise - Full Employee Directory"
            title_cell.font = font_title
            title_cell.fill = fill_primary
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws1.row_dimensions[1].height = 40

            # Headers
            headers1 = [
                "Employee ID",
                "Name",
                "Department",
                "Designation",
                "Monthly Salary",
                "Joining Date",
                "Performance Score",
                "Annual Salary",
                "Annual Bonus",
                "Annual Tax",
                "Net Annual Salary",
            ]

            ws1.append([])  # Blank row
            ws1.append(headers1)
            ws1.row_dimensions[3].height = 25

            # Format header row
            for col_idx in range(1, len(headers1) + 1):
                cell = ws1.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_primary
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_cell

            # Write employee rows
            for emp in employees:
                # Compile row values (including financial calculations)
                row_data = [
                    emp.employee_id,
                    emp.name,
                    emp.department,
                    emp.designation,
                    emp.salary,
                    emp.joining_date,
                    emp.performance_score,
                    emp.annual_salary(),
                    emp.calculate_bonus(),
                    emp.calculate_tax(),
                    emp.net_salary(),
                ]
                ws1.append(row_data)
                curr_row = ws1.max_row
                ws1.row_dimensions[curr_row].height = 18

                # Format data cells
                for col_idx in range(1, len(row_data) + 1):
                    c = ws1.cell(row=curr_row, column=col_idx)
                    c.font = font_data
                    c.border = border_cell

                    # Number format conversions
                    if col_idx in [1, 7]:  # ID, Performance
                        c.alignment = Alignment(horizontal="center")
                    elif col_idx in [5, 8, 9, 10, 11]:  # Currency fields
                        c.number_format = "₹#,##0.00"
                        c.alignment = Alignment(horizontal="right")
                    elif col_idx == 6:  # Date
                        c.alignment = Alignment(horizontal="center")

            # Auto-fit columns to avoid truncated fields
            for col in ws1.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                # Skip title row cell length to avoid bloating column width
                for cell in col[2:]:
                    if cell.value:
                        # Add commas/decimals buffer
                        max_len = max(max_len, len(str(cell.value)))
                ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

            # ==============================================================================
            # SHEET 2: DEPARTMENT AGGREGATES
            # ==============================================================================
            ws2 = wb.create_sheet(title="Department Aggregates")
            ws2.views.sheetView[0].showGridLines = True

            ws2.merge_cells("A1:E1")
            title_cell2 = ws2["A1"]
            title_cell2.value = "RetailMax HR Department Aggregates"
            title_cell2.font = font_title
            title_cell2.fill = fill_primary
            title_cell2.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[1].height = 40

            headers2 = [
                "Department",
                "Headcount",
                "Total Salary Expense",
                "Average Salary",
                "Max Salary",
            ]
            ws2.append([])
            ws2.append(headers2)
            ws2.row_dimensions[3].height = 25

            for col_idx in range(1, len(headers2) + 1):
                cell = ws2.cell(row=3, column=col_idx)
                cell.font = font_header
                cell.fill = fill_primary
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border_cell

            # Populate metrics
            for dept in dept_summary:
                row_data = [
                    dept["Department"],
                    dept["Headcount"],
                    dept["TotalSalary"],
                    dept["AverageSalary"],
                    dept["MaxSalary"],
                ]
                ws2.append(row_data)
                curr_row = ws2.max_row
                ws2.row_dimensions[curr_row].height = 20

                # Formats
                for col_idx in range(1, len(row_data) + 1):
                    c = ws2.cell(row=curr_row, column=col_idx)
                    c.font = font_data
                    c.border = border_cell
                    if col_idx == 2:
                        c.alignment = Alignment(horizontal="center")
                    elif col_idx in [3, 4, 5]:
                        c.number_format = "₹#,##0.00"
                        c.alignment = Alignment(horizontal="right")

            # Totals summary row
            total_row_idx = ws2.max_row + 1
            ws2.cell(row=total_row_idx, column=1, value="Grand Total").font = font_summary
            ws2.cell(row=total_row_idx, column=1).border = border_summary
            ws2.cell(row=total_row_idx, column=1).fill = fill_accent

            # Formula calculations
            ws2.cell(row=total_row_idx, column=2, value=f"=SUM(B4:B{total_row_idx-1})").font = (
                font_summary
            )
            ws2.cell(row=total_row_idx, column=2).alignment = Alignment(horizontal="center")
            ws2.cell(row=total_row_idx, column=2).border = border_summary
            ws2.cell(row=total_row_idx, column=2).fill = fill_accent

            ws2.cell(row=total_row_idx, column=3, value=f"=SUM(C4:C{total_row_idx-1})").font = (
                font_summary
            )
            ws2.cell(row=total_row_idx, column=3).number_format = "₹#,##0.00"
            ws2.cell(row=total_row_idx, column=3).alignment = Alignment(horizontal="right")
            ws2.cell(row=total_row_idx, column=3).border = border_summary
            ws2.cell(row=total_row_idx, column=3).fill = fill_accent

            ws2.cell(row=total_row_idx, column=4, value=f"=AVERAGE(D4:D{total_row_idx-1})").font = (
                font_summary
            )
            ws2.cell(row=total_row_idx, column=4).number_format = "₹#,##0.00"
            ws2.cell(row=total_row_idx, column=4).alignment = Alignment(horizontal="right")
            ws2.cell(row=total_row_idx, column=4).border = border_summary
            ws2.cell(row=total_row_idx, column=4).fill = fill_accent

            ws2.cell(row=total_row_idx, column=5, value=f"=MAX(E4:E{total_row_idx-1})").font = (
                font_summary
            )
            ws2.cell(row=total_row_idx, column=5).number_format = "₹#,##0.00"
            ws2.cell(row=total_row_idx, column=5).alignment = Alignment(horizontal="right")
            ws2.cell(row=total_row_idx, column=5).border = border_summary
            ws2.cell(row=total_row_idx, column=5).fill = fill_accent

            ws2.row_dimensions[total_row_idx].height = 22

            # Auto-fit sheet 2 columns
            for col in ws2.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col[2:]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws2.column_dimensions[col_letter].width = max(max_len + 5, 15)

            wb.save(excel_path)
            logger.info(f"Excel workbook report generated at: {excel_path}")
            return excel_path

        except Exception as e:
            raise AnalyticsError(f"Failed to generate Excel report using OpenPyXL: {e}") from e


# ==============================================================================
# INTERVIEW NOTES & SPREADSHEET ARCHITECTURES:
#
# Q1: Why do business executives prefer Excel (.xlsx) over raw CSV files?
#     CSV is a flat text format containing only comma-separated text values,
#     supporting no styling, formulas, column sizing, borders, or multi-sheet layouts.
#     Excel Workbooks (.xlsx) are zipped XML structures supporting rich cell formatting,
#     number layouts (like currency and date symbols), cell merging, embedded charts,
#     and mathematical formulas (like '=SUM(B4:B9)'), which enable self-calculating reports.
#
# Q2: How does OpenPyXL build Excel files?
#     OpenPyXL generates worksheets in memory, mapping grids into objects.
#     Styles (Fonts, Fills, Borders) must be instantiated as separate objects
#     and assigned individually to cells. For large datasets, using row appends
#     ('ws.append') and post-styling is memory efficient.
#
# Q3: Why is 'ws.views.sheetView[0].showGridLines = True' needed?
#     By default, when you apply cell styling fills (like PatternFill) to cell ranges,
#     some spreadsheet software (like MS Excel) hides the default gridlines.
#     Setting showGridLines explicitly to True guarantees that cells retain visible
#     border grids, maintaining clean visual alignments.
# ==============================================================================
